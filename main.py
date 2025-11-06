#! /usr/bin/env python3
# encoding: utf-8

import subprocess
import arch_platforms

import psycopg2
from rich.console import Console
from rich.table import Table
from rich.traceback import install
from rich.logging import RichHandler
from setproctitle import setproctitle
from datetime import datetime

import logging
import argparse
import tomllib
import time
import json,pickle
import faker
import zipfile,lzma,tarfile
import shutil,psutil
import os,sys,signal
import requests
import humanfriendly
from collections import namedtuple
from pathlib import Path,PurePosixPath
from io import BytesIO
from openpyxl import load_workbook
import zstandard as zstd
from paramiko import SSHClient
from psycopg2 import sql
from psycopg2.extras import Json
from psycopg2.pool import ThreadedConnectionPool
from concurrent.futures import ThreadPoolExecutor
from queue import Queue,Empty
from pySmartDL import SmartDL



setproctitle('mrcb')    # 设置mrcb的进程名称

#install(show_locals=True)
console = Console(color_system='256',file=sys.stdout)

logging.basicConfig(
    level=logging.INFO,
    format='%(message)s',
    handlers=[RichHandler()]
)


cpu_count = os.cpu_count()
pgsql_pool = ThreadedConnectionPool(
    minconn=1,maxconn=cpu_count,
    host='localhost',
    port=5432,
    user='postgres',
    password='postgres',
    dbname='mugen_run_clean_batch',
)


# mugen测试用例描述
mugen_test = namedtuple(
    typename='mugen_test',
    field_names=[
        'TestSuite',    # 所属测试套
        'TestCase'      # 测试用例名
    ]
)
mugen_tests:Queue = Queue(maxsize=cpu_count * 2)      # 存放所有测试用例描述的列表
target_mugen_tail_object:object = object()             # 队列尾标志
# 把所有待测试项目全部放进队列里
def put_mugen_test_to_queue(mugen_test_list:list):
    for mugen_test in mugen_test_list:
        mugen_tests.put(mugen_test,block=True)
    mugen_tests.put(target_mugen_tail_object,block=True)




# mrcb存放资源的临时目录
mrcb_dir = Path('.')
current_strftime = datetime.now().strftime('%Y%m%d_%H%M%S')
mrcb_work_dir = mrcb_dir / f"workdir_{current_strftime}" # 本次运行mrcb创建的工作目录
mrcb_firmware_dir = mrcb_work_dir / 'firmware'   # 存放固件
mrcb_mugen_dir = mrcb_work_dir / 'mugen'         # 存放mugen项目
mrcb_runtime_dir = mrcb_work_dir / 'runtime'
mrcb_runtime_default_dir = mrcb_runtime_dir / 'default'

single_machine_tests = []
multi_machine_tests = []


# 当前项目支持的架构和启动平台
support_platforms = ('UEFI','UBOOT','PENGLAI')  # 支持的启动引导方式(固件)
support_archs = ('X86','ARM','RISC-V')          # 支持的指令集架构


# 预制的https请求头
faker = faker.Faker()
headers = {
    'Accept': 'image/avif,image/webp,image/png,image/svg+xml,image/*;q=0.8,*/*;q=0.5',
    'Accept-Encoding': 'gzip, deflate, br, zstd',
    'User-Agent': faker.user_agent(),
    'Referer': 'https://gitee.com/April_Zhao/mugen_run_clean_batch',
}




# 只提取Toml参数不作处理
def parse_config() -> dict:
    """
    1. 命令行交互
    2. 读取和解析Toml配置文件
    :return: config字典,所有输入的参数转换为Python对象
    """
    parser = argparse.ArgumentParser(
        description="mrcb - Batch run mugen tests in a clean environment\n"
                    "把mugen批量地运行在干净的系统环境"
    )
    parser.add_argument(
        "--config", "-c",
        type=str,
        default="mrcb_config.toml",
        help="指定mrcb所需的Toml格式配置文件"
    )

    mrcb_config_file = parser.parse_args().config
    try:
        config = tomllib.loads(open(mrcb_config_file).read())
    except FileNotFoundError:
        console.print(f"您指定的文件{mrcb_config_file}不存在,请检查文件或目录名是否正确")
        sys.exit(1)

    # 判断platform是否合法
    global platform,arch
    platform,arch = config['platform'],config['arch']
    if platform not in support_platforms:
        console.print(f'您输入的openEuler qemu镜像启动平台{platform}值不合法,合法值必须为{support_platforms}其中之一')
        console.print('如有新平台需要接入请给mrcb项目提交issue')
        sys.exit(1)
    print(config)
    return config



# 校验输入的url是否是可访问的
def check_url(url: str) -> bool:
    try:
        return requests.head(url=url,allow_redirects=True,timeout=10).ok
    except requests.RequestException:
        return False


# 对选项参数作处理
def check_config(config:dict)->dict:
    """
    检查用户输入的Toml内的值是否合法
    根据配置文件的值制作参数列表
    :param platform:
    :param config:
    :return:
    """
    arch = config.get('arch',None)
    if arch is None or arch not in support_archs:
        console.print(f'您输入的arch字段有误,请确认值为下列之一:{support_archs}')
        sys.exit(1)
    result = {}

    drive_type = config.get('drive_type',None)
    if drive_type is None:
        print("drive_type字段为空,请检查输入的Toml")
    result['drive_type'] = drive_type

    compress_format = config.get('compress_format',None)
    if compress_format is None:
        print("compress_format字段为空,则输入的url为不压缩的镜像!!!")
    result['compress_format'] = compress_format

    mrcb_runtime_dir.mkdir(exist_ok=True);mrcb_firmware_dir.mkdir(exist_ok=True)

    if arch == 'RISC-V':
        if platform == "UEFI":
            drive_url:str = config.get('drive_url','')
            if drive_url == '' or check_url(drive_url) is False:
                console.print(f'您输入的drive_url字段url无法访问,请检查')
                #sys.exit(1)
            result['drive_name'] = PurePosixPath(drive_url).name


            download_drive_file:SmartDL = SmartDL(
                urls = [
                    drive_url
                ],
                dest = str(mrcb_runtime_default_dir / result['drive_name']),
                threads = min(cpu_count,32),
                timeout=10,
                progress_bar=True,
            )
            download_drive_file.start(blocking=False)

            VIRT_CODE = config.get('VIRT_CODE','')
            if VIRT_CODE == '' or check_url(VIRT_CODE) is False:
                console.print(f"您输入的VIRT_CODE字段url无法访问,请检查")
                sys.exit(1)
            result['VIRT_CODE'] = VIRT_CODE
            result['VIRT_CODE_FILE'] = mrcb_firmware_dir / PurePosixPath(VIRT_CODE).name
            download_VIRT_CODE_file:SmartDL = SmartDL(
                urls = [VIRT_CODE],
                dest = str(result['VIRT_CODE_FILE']),
                threads = min(cpu_count,32),
                timeout=10,
                progress_bar=True,
            )
            download_VIRT_CODE_file.start(blocking=True)


            VIRT_VARS = config.get('VIRT_VARS','')
            if VIRT_VARS == '' or check_url(VIRT_VARS) is False:
                console.print(f"您输入的VIRT_VARS字段url无法访问,请检查")
                sys.exit(1)
            result['VIRT_VARS'] = VIRT_VARS
            result['VIRT_VARS_FILE'] = mrcb_firmware_dir / PurePosixPath(VIRT_VARS).name
            download_VIRT_VARS_file:SmartDL = SmartDL(
                urls = [VIRT_VARS],
                dest = str(result['VIRT_VARS_FILE']),
                threads = min(cpu_count,32),
                timeout=10,
                progress_bar=True,
            )
            download_VIRT_VARS_file.start(blocking=True)
            download_drive_file.wait()

        elif platform == "UBOOT":
            drive_url:str = config.get('drive_url','')
            if drive_url == '' or check_url(drive_url) is False:
                console.print(f'您输入的drive_url字段url无法访问,请检查')
                sys.exit(1)
            result['drive_name'] = PurePosixPath(drive_url).name

            download_drive_file:SmartDL = SmartDL(
                urls = [
                    drive_url
                ],
                dest = str(mrcb_runtime_default_dir / result['drive_name']),
                threads = min(cpu_count,32),
                timeout=10,
                progress_bar=True,
            )
            download_drive_file.start(blocking=False)

            uboot_bin:str = config.get('uboot_bin','')
            if uboot_bin == '' or check_url(uboot_bin) is False:
                print(f'您输入的uboot_bin字段为空,请检查Toml文件')
                sys.exit(1)
            result['UBOOT_BIN_FILE'] = mrcb_firmware_dir / PurePosixPath(uboot_bin).name
            download_uboot_bin_file:SmartDL = SmartDL(
                urls = [uboot_bin],
                dest = str(result['UBOOT_BIN_FILE']),
                threads = min(cpu_count,32),
                timeout=10,
                progress_bar=True,
            )
            download_uboot_bin_file.start(blocking=True)
            download_drive_file.wait()

        elif platform == "PENGLAI":
            pass
    input_excel = config.get('input_excel','')
    if input_excel == '':
        console.print(f"input_excel字段不可以为空")
        sys.exit(1)
    from_to = config.get('from_to',[])
    if from_to is []:
        print("from_to字段填写错误，请查看README中相关描述")
        sys.exit(1)
    result['from_to'] = from_to
    result['input_excel'] = input_excel
    return result

    # # 检测完成后建立数据表,将运行信息登记
    # with pgsql_pool.getconn() as conn:
    #     with conn.cursor() as cursor:
    #         cursor.execute("""
    #         intert into mugen_run_clean_batch values(
    #
    #         );
    #     """)
        


def get_analysis_mugen():
    """
    初始化正式运行前所需的本地环境
    :return:
    """

    if mrcb_work_dir.exists():
        shutil.rmtree(mrcb_work_dir)
    mrcb_work_dir.mkdir();mrcb_firmware_dir.mkdir()
    # 下载所有所需文件到tmp目录
    try:    # 获取mugen
        subprocess.run(
            args="git clone https://gitee.com/openeuler/mugen.git --depth=1",
            cwd=mrcb_work_dir,
            check=True,shell=True,
        )
    except subprocess.CalledProcessError as e:
        console.print(f"git clone失败,{e}")
        sys.exit(1)
    # 获取所有有可能用到的json文件的文件名
    global mugen_suite_jsons, mugen_cli_test_jsons, mugen_doc_test_jsons, mugen_fs_test_jsons, mugen_network_test_jsons, mugen_service_jsons, mugen_smoke_test_jsons, mugen_system_integration_jsons
    mugen_suite_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases'))[0][2]
    mugen_cli_test_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'cli-test'))[0][2]
    mugen_doc_test_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'doc-test'))[0][2]
    mugen_fs_test_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'fs-test'))[0][2]
    mugen_network_test_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'network_test'))[0][2]
    mugen_service_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'service'))[0][2]
    mugen_smoke_test_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'smoke-test'))[0][2]
    mugen_system_integration_jsons = list(os.walk(mrcb_mugen_dir / 'suite2cases/mugen_baseline_json' / 'system-integration'))[0][2]




def input_from_excel():
    input_excel_file = config.get('input_excel')
    wb = load_workbook(input_excel_file,read_only=True)
    ws = wb.active
    from_to = config.get('from_to',[])

    all_mugen_tests = []
    # 获取所有需要测试的mugen测试用例名
    for i in range(from_to[0],from_to[1]+1):
        each_mugen_test = mugen_test(
            TestSuite = ws.cell(row=i,column=1).value,
            TestCase = ws.cell(row=i,column=2).value,
        )
        all_mugen_tests.append(each_mugen_test)

    conn = pgsql_pool.getconn()
    cursor = conn.cursor()
    # 校验测试用例的合法性
    for TestSuite,TestCase in all_mugen_tests:
        # suite2case
        if TestSuite + '.json' in mugen_suite_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / f'{TestSuite}.json','r',encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                #print(TestSuite_json)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name,testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                    "insert into {} (testsuite,testcase,desc_json)"
                    "values(%s,%s,%s) RETURNING id;"
                ).format(sql.Identifier(f"workdir_{current_strftime}")),(TestSuite,TestCase,Json(TestSuite_json)))
                conn.commit()

        # baseline test
        elif TestSuite + '.json' in mugen_cli_test_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'cli-test' / f'{TestSuite}.json', 'r', encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()

        elif TestSuite + '.json' in mugen_doc_test_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'doc-test' / f'{TestSuite}.json', 'r',
                      encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        elif TestSuite + '.json' in mugen_fs_test_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'fs-test' / f'{TestSuite}.json', 'r',
                      encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        elif TestSuite + '.json' in mugen_network_test_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'network_test' / f'{TestSuite}.json', 'r',
                      encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        elif TestSuite + '.json' in mugen_service_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'service' / f'{TestSuite}.json', 'r',
                      encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        elif TestSuite + '.json' in mugen_smoke_test_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'smoke-test' / f'{TestSuite}.json', 'r', encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        elif TestSuite + '.json' in mugen_system_integration_jsons:
            with open(mrcb_mugen_dir / 'suite2cases' / 'mugen_baseline_json' / 'system-integration' / f'{TestSuite}.json', 'r',
                      encoding='utf-8') as f:
                TestSuite_json = json.load(f)
                if TestCase not in (testcase for each in TestSuite_json['cases'] for name, testcase in each.items()):
                    print(f"{TestSuite}中不含有{TestCase},请仔细检查excel文件!")
                # 将获取到的信息写入数据库
                cursor.execute(
                    sql.SQL(
                        "insert into {} (testsuite,testcase,desc_json)"
                        "values(%s,%s,%s) RETURNING id;"
                    ).format(sql.Identifier(f"workdir_{current_strftime}")),
                    (TestSuite, TestCase, Json(TestSuite_json)))
                conn.commit()
        else:
            print(f"{TestSuite}-{TestCase}测试未在mugen中找到,请重新检查输入的excel文件.")
    cursor.close()
    pgsql_pool.putconn(conn)



def init_postgresql():
    with pgsql_pool.getconn() as conn:
        cursor = conn.cursor()
        cursor.execute(
        sql.SQL (
            "CREATE TABLE {} ("
            "id integer GENERATED BY DEFAULT AS IDENTITY PRIMARY KEY,"
            "testsuite varchar(30) NOT NULL,"
            "testcase varchar(100) NOT NULL,"
            "desc_json json NOT NULL,"
            "state boolean NOT NULL default FALSE,"
            "start_time timestamptz,"
            "end_time timestamptz,"
            "check_result char(7),"
            "output_log text,"
            "failure_reason text)"
        ).format(sql.Identifier(f"workdir_{current_strftime}")))
        conn.commit()
        cursor.close()


def init_internet_gateway():
    # 若br0不存在
    br0 = subprocess.run(
        args = "brctl show br0",
        shell = True,
    )
    if br0.returncode != 0:
        # 宿主机创建网桥,并给网桥配置IP
        br0 = subprocess.run(
            args = "brctl addbr br0 && "
                   "ip link set br0 up && "
                   "ip addr add 10.0.0.1/24 dev br0",
            shell = True,
            stdout = subprocess.PIPE,
            stderr = subprocess.PIPE,
        )
        if br0.returncode != 0:
            print(f"初始化网关失败.报错信息:{br0.stderr.decode('utf-8')}")
            print("详细请参考:https://github.com/openeuler-riscv/oerv-qa/blob/main/docs/mugen/Mugen%E6%B5%8B%E8%AF%95Lesson%20Learn.md")
            #sys.exit(1)
    else:
        set_br0 = subprocess.run(
            args = "ip link set br0 up && "
                    "ip addr add 10.0.0.1/24 dev br0",
                shell = True,
                stdout = subprocess.PIPE,
                stderr = subprocess.PIPE,
            )


    # 宿主机添加虚拟网卡(有几台vm就需要几个tap)全都挂到同一个br0上面
    try:
        for i in range(1,cpu_count+1):
            subprocess.run(
                args = f"ip tuntap add tap{i} mode tap &&"
                       f"brctl addif br0 tap{i} &&"
                       f"ip link set tap{i} up",
                shell = True,
                check = True,
                stdout = subprocess.PIPE,
                stderr = subprocess.PIPE,
            )
    except subprocess.CalledProcessError as e:
        print(f"创建虚拟网卡失败.报错信息:{e.stderr.decode('utf-8')}")



# 制作镜像模板
def make_template_image():
    drive_name = config.get('drive_name')
    # 依据引导选项和指令集做判断
    if arch == 'RISC-V':
        if platform == 'UEFI':
            arch_platforms.RISC_V_UEFI.make_openEuler_image(
                **{'default_workdir':mrcb_runtime_default_dir,
                   'VIRT_CODE_FILE':config['VIRT_CODE_FILE'],
                   'VIRT_VARS_FILE':config['VIRT_VARS_FILE'],
                   'DRIVE_FILE':drive_name,'DRIVE_TYPE':config['drive_type'],
                   'compress_format':config['compress_format'],
                   'DEVICE_TYPE':config['device_type'],
                   'mugen_dir':mrcb_mugen_dir
                })
        elif platform == 'UBOOT':
            arch_platforms.RISC_V_UBOOT.make_openEuler_image(
                **{
                    'default_workdir':mrcb_runtime_default_dir,
                    'UBOOT_BIN_FILE':config['UBOOT_BIN_FILE'],
                    'DRIVE_FILE': drive_name, 'DRIVE_TYPE': config['drive_type'],
                    'compress_format': config['compress_format'],
                    'mugen_dir': mrcb_mugen_dir
                }
            )



if __name__ == "__main__":
    start_time = time.time()

    # 先初始化mugen
    get_analysis_mugen()
    config:dict = parse_config()
    config:dict = check_config(config)
    init_postgresql()
    input_from_excel()
    init_internet_gateway()
    make_template_image()





    # 正式开始测试
    # with ThreadPoolExecutor(max_workers=cpu_count) as executor:
    #     executor.submit(put_mugen_test_to_queue)


    end_time = time.time()
    console.print(f"mrcb运行结束,本次运行总耗时{humanfriendly.format_timespan(end_time - start_time)}")